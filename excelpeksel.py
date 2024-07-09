from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl.styles import PatternFill, Font, Border, Side, PatternFill
from datetime import datetime, timedelta

def lunch_time(workbook):
    # Определяем рабочую таблицу
    worksheet = workbook.worksheets[1]
    for row in range(1, worksheet.max_row):
        interval = str(worksheet.cell(row=row, column=5).value)
        # Проеверяем наличие символа '-' в строке
        if '-' in interval:
            # Разбиваем строку на начальное и конечное время
            start_time, end_time = interval.split('-')
            try:
                # Перобразуем строки в формат времени
                start_time = datetime.strptime(start_time, "%H:%M")
                end_time = datetime.strptime(end_time, "%H:%M")
                # Вычисляем разницу
                time_delta = end_time - start_time
                # Делаем проверку
                if time_delta > timedelta(minutes=30):
                    worksheet.cell(row=row, column=5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    workbook.save("NewBook.xlsm")

                # Выводим результат
                # !!! Нужно понимать куда вставлять данные и что с ними делать !!!
                print(f"Pavza malica {row}: {time_delta}")
            except ValueError as error:
                print(f"No values Error: {row}: {error}")
        else:
            print(f"No values Error: {row}")

def total_line(workbook):
    # Устанавливаем рабочий лист для работы
    worksheet = workbook.worksheets[1]
    # Находим первую строку с днем недели
    for row in range(1, worksheet.max_row + 1):
        week_days = worksheet.cell(row=row, column=2).value
        # Если в списке будет 'nedelja' то вставляем строку после этого поля
        if week_days == 'nedelja':
            worksheet.insert_rows(row + 1, amount=1)
            # Записываем "Skupaj" в новую ячейку
            worksheet.cell(row=row + 1, column=1, value="Skupaj").font = Font(bold=True)
            # Устанавливаем рамку для новой строки
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            for cell in worksheet[row + 1]:
                cell.border = border
    # Сохраняем excel file        
    workbook.save('newfile.xlsm')

def holidays(workbook, holidates):
    worksheet = workbook.worksheets[1]
    # Форматируем праздники в формат datetime
    dates = [datetime.strptime(date, '%d.%m.%Y').date() for date in holidates]
    # Итерируем и находим колонку с нужными нам значениями
    for row in range(1, worksheet.max_row +1):
        days = worksheet.cell(row=row, column=1).value
        # Проверяем даты на соответствие примеру
        if days and isinstance(days, datetime):
            days_dates = days.date()
            # Проверяем даты на соответствие праздничных дат
            if days_dates in dates:
                for cell in worksheet[row]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    workbook.save("Colorexcel.xlsm")

def process_file(filename):
    try:
        workbook = load_workbook(filename)
        lunch_time(workbook)
        total_line(workbook)
        holidays(workbook, [
        '01.01.2023', '02.01.2023', '08.02.2023', 
        '27.04.2023', '01.05.2023', '08.06.2023', 
        '25.06.2023', '14.08.2023', '17.08.2023', 
        '15.09.2023', '23.09.2023', '25.10.2023', 
        '31.10.2023', '01.11.2023', '23.11.2023', 
        '25.12.2023', '26.12.2023'
        ])
        messagebox.showinfo("Zakljuceno uspešno!")
    except Exception as e:
        messagebox.showerror("Napaka", f"Napaka: {str(e)}")

def select_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Excel Macros-Enable", "*.xlsm")])
    if filename:
        process_file(filename)

# Создаем основное окно Tkinter
root = tk.Tk()
root.title("Excel Processor")

# Создаем кнопку для выбора файла
select_button = tk.Button(root, text="Izberi datoteko", command=select_file)
select_button.pack(pady=10)
root.mainloop()
