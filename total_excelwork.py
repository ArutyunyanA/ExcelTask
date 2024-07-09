from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime, timedelta

def lunch_time(workbook, sheet_index):
    worksheet = workbook.worksheets[sheet_index - 1]
    for row in range(1, worksheet.max_row + 1):
        interval = str(worksheet.cell(row=row, column=5).value)
        if '-' in interval:
            # Split the interval only if it contains exactly two components
            if interval.count('-') == 1:
                start_time, end_time = interval.split('-')
                try:
                    start_time = datetime.strptime(start_time, "%H:%M")
                    end_time = datetime.strptime(end_time, "%H:%M")
                    time_delta = end_time - start_time
                    if time_delta > timedelta(minutes=30):
                        worksheet.cell(row=row, column=5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                except ValueError as error:
                    print(f"Error processing row {row}: {error}")

def holidays(workbook, sheet_index, holidates):
    worksheet = workbook.worksheets[sheet_index]
    dates = [datetime.strptime(date, '%d.%m.%Y').date() for date in holidates]
    for row in range(1, worksheet.max_row + 1):
        days = worksheet.cell(row=row, column=1).value
        if days and isinstance(days, datetime):
            days_dates = days.date()
            if days_dates in dates:
                for cell in worksheet[row]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def total_line(workbook, sheet_index):
    worksheet = workbook.worksheets[sheet_index]
    for row in range(1, worksheet.max_row + 1):
        week_days = worksheet.cell(row=row, column=2).value
        if week_days == 'nedelja':
            worksheet.insert_rows(row + 1, amount=1)
            worksheet.cell(row=row + 1, column=1, value="Skupaj").font = Font(bold=True)
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            for cell in worksheet[row + 1]:
                cell.border = border

if __name__ == '__main__':
    workbook = load_workbook('Delovniki - DECEMBER 2023.xlsm')
    for sheet_index in range(1, 5):  # Adjusted for four sheets starting from index 1
        lunch_time(workbook, sheet_index)
        holidays(workbook, sheet_index, [
            '01.01.2023', '02.01.2023', '08.02.2023', 
            '27.04.2023', '01.05.2023', '08.06.2023', 
            '25.06.2023', '14.08.2023', '17.08.2023', 
            '15.09.2023', '23.09.2023', '25.10.2023', 
            '31.10.2023', '01.11.2023', '23.11.2023', 
            '25.12.2023', '26.12.2023'
        ])
        total_line(workbook, sheet_index)
    workbook.save("Colorexcel.xlsm")
